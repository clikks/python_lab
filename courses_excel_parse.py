#!/usr/bin/env python3
# -*- coding: utf-8 -*-
__author__ = 'lux'

from multiprocessing import Pool
import datetime
import os

from openpyxl import load_workbook, Workbook


file = os.path.join(os.getcwd(), 'courses.xlsx')


def combine():
    start = datetime.datetime.now()
    print("func 'combine' is runing")
    inwb = load_workbook(file)
    student = inwb['students']
    study_time = inwb['time']
    inwb.create_sheet('combine')
    combine = inwb['combine']
    time_dict = dict()
    for index, row in enumerate(study_time.iter_rows()):
        if index == 0:
            title = [row[-1].value]
        else:
            time_dict[row[1].value] = row[-1].value
        # course = 'B' + str(row)
        # time = 'C' + str(row)
        # time_dict[study_time[course].value] = study_time[time].value
    for index, row in enumerate(student.iter_rows()):
        if index == 0:
            combine_title = [i.value for i in row] + title
            combine.append(combine_title)
        else:
            cell_content = [i.value for i in row] + [time_dict.get(row[1].value)]
            combine.append(cell_content)
            # for seq, data in enumerate(row):
            #     combine.cell(row=index+1, column=seq+1, value=data.value)
            #     time_value = time_dict.get(data.value)
            #     if seq == 0:
            #         combine.cell(row=index+1, column=student.max_column+1, value=study_time['C1'].value)
            #     else:
            #         combine.cell(row=index+1, column=student.max_column+1, value=time_value)
    inwb.save('courses.xlsx')
    end = datetime.datetime.now()
    print("func 'combine' finished, used time {}s".format((end-start).seconds))


def generate_wb(year, row_title, data_dict):
    child_start = datetime.datetime.now()
    print('Run child process {}'.format(os.getpid()))
    newwb = Workbook()
    sheet = newwb.active
    sheet.title = str(year)
    sheet.append(row_title)
    for date in data_dict.keys():
        if date.year == year:
            cells = [date] + data_dict.get(date)
            sheet.append(cells)
    filename = str(year) + '.xlsx'
    newwb.save(filename)
    child_end = datetime.datetime.now()
    print('Process {} finished, used time {}s'.format(os.getpid(), (child_end - child_start).seconds))


def split():
    start = datetime.datetime.now()
    print("func 'split' is running")
    inwb = load_workbook(file)
    combine = inwb['combine']
    sheet_dict = dict()
    for index, row in enumerate(combine.iter_rows()):
        if index == 0:
            row_title = [i.value for i in row]
        else:
            sheet_dict[row[0].value] = [i.value for i in row[1:]]
    years = set(key.year for key in sheet_dict.keys())
    pool = Pool(processes=4)
    for year in years:
        pool.apply_async(generate_wb, (year, row_title, sheet_dict))
    pool.close()
    pool.join()
    end = datetime.datetime.now()
    print("func 'split' is finished, used time {}s".format((end-start).seconds))


if __name__ == '__main__':
    print('Program run...')
    starttime = datetime.datetime.now()
    combine()
    split()
    endtime = datetime.datetime.now()
    print('Program is finished runing. used time {}s'.format((endtime-starttime).seconds))